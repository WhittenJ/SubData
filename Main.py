import openpyxl
import pprint
import re
import tkinter as tk
from tkinter import filedialog


def get_excel():
    # Set up File Selection
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename()

    return file_path


def import_excel(file_path):
    # Set up Excel using OpenPyXL
    print('Reading in Sub data...')

    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sub Loot']
    sub_data = []

    print('Read in Sub Data...')
    for row in range(2, sheet.max_row + 1):
        # Each row in the spreadsheet has data
        # Row A = Sector (URN kind of)
        # Row B = Workshop Materials 1-8
        # Row C = Crafting Materials 1-8
        # Row D = Materia 1-8 (not useful)
        # Row E = Shard Crystal Cluster 1-6
        # Row F = Rare Items 1-8

        if row < 35:
            map_name = 'Deep-sea Site'
        else:
            map_name = 'Sea of Ash'
        sector = str(sheet['A' + str(row)].value).splitlines()
        wsm = str(sheet['B' + str(row)].value).splitlines()
        cm = str(sheet['C' + str(row)].value).splitlines()
        materia = str(sheet['D' + str(row)].value).splitlines()
        scc = str(sheet['E' + str(row)].value).splitlines()
        rare_items = str(sheet['F' + str(row)].value).splitlines()

        if sector[0] == 'None' or sector[0] == 'Sea of Ash':
            continue
        else:
            sub_data.append([map_name, sector, wsm, cm, materia, scc, rare_items])

    return sub_data


def export_excel(sub_data):
    # Create blank Workbook
    wb = openpyxl.Workbook()

    # Get active sheet
    sheet = wb.active

    # Sub_Data[0] = Sector (URN kind of)
    # Sub_Data[1][0-7] = Workshop Materials 1-8
    # Sub_Data[2][0-7] = Crafting Materials 1-8
    # Sub_Data[3][0-7] = Materia 1-8 (not useful)
    # Sub_Data[4][0-5] = Shard Crystal Cluster 1-6
    # Sub_Data[5][0-7] = Rare Items 1-8
    sheet['A1'].value = 'Map'
    sheet['B1'].value = 'Sector'
    sheet['C1'].value = 'Category'
    # Categories = Workshop, Crafting, Materia, Shard Crystal Cluster, Rare
    sheet['D1'].value = 'Item'

    for read, data in enumerate(sub_data):
        row = sheet.max_row + 1

        # Map
        map_name = data[0]
        # Sector
        try:
            data[1][2]
        except TypeError:
            continue
        if '★' not in data[1][2]:
            sector = sheet['B' + str(row)].value = data[1][1].strip() + ' ' + data[1][2].strip()
        else:
            sector = sheet['B' + str(row)].value = data[1][1].strip()

        new_row = 99

        # Workshop Materials
        for item in data[2]:
            if new_row == 99:
                new_row = sheet.max_row
            else:
                new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = map_name
            sheet['B' + str(new_row)].value = sector
            sheet['C' + str(new_row)].value = 'Workshop Materials'
            wsm = sheet.cell(row=new_row, column=4)
            wsm.value = re.sub(r" ?\([^a-z)]+\)", "", item)
        # Crafting Materials
        for item in data[3]:
            new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = map_name
            sheet['B' + str(new_row)].value = sector
            sheet['C' + str(new_row)].value = 'Crafting Materials'
            cm = sheet.cell(row=new_row, column=4)
            cm.value = item
        ''' # Materia
        for item in data[4]:
            new_row = sheet.max_row + 1
            sheet['B' + str(new_row)].value = 'Materia'
            materia = sheet.cell(row=new_row, column=4)
            materia.value = item '''
        # Shard Crystal Cluster
        for item in data[5]:
            new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = map_name
            sheet['B' + str(new_row)].value = sector
            sheet['C' + str(new_row)].value = 'Shard Crystal Cluster'
            scc = sheet.cell(row=new_row, column=4)
            scc.value = item
        # Rare Items
        for item in data[6]:
            new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = map_name
            sheet['B' + str(new_row)].value = sector
            sheet['C' + str(new_row)].value = 'Rare Items'
            ri = sheet.cell(row=new_row, column=4)
            ri.value = re.sub(r" ?\([^a-z)]+\)", "", item)

    wb.save('C:\\Users\\jwhitten\\Documents\\Python\\Sub Data\\ExportData.xlsx')


def main():
    file_path = get_excel()
    sub_data = import_excel(file_path)

    # for x in sub_data:
    # print(x)

    # TODO Export Excel data
    export_excel(sub_data)


if __name__ == "__main__":
    main()
