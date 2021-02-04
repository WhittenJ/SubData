import openpyxl
import pprint
import re
import tkinter as tk
from tkinter import filedialog


def get_excel():
    # Set up File Selection
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename()

    return file_path


def import_excel(file_path):
    # Set up Excel using OpenPyXL
    print('Reading in Airship data...')

    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Airship Loot']
    airship_data = []

    print('Read in Airship Data...')
    for row in range(2, sheet.max_row + 1):
        # Each row in the spreadsheet has data
        # Row A = Sector (URN kind of)
        # Row B = Workshop Materials 1-8
        # Row C = Crafting Materials 1-8
        # Row D = Materia 1-8 (not useful)
        # Row E = Shard Crystal Cluster 1-6
        # Row F = Rare Items 1-8

        sector = str(sheet['A' + str(row)].value).splitlines()
        wsm = str(sheet['B' + str(row)].value).splitlines()
        cm = str(sheet['C' + str(row)].value).splitlines()
        materia = str(sheet['D' + str(row)].value).splitlines()
        scc = str(sheet['E' + str(row)].value).splitlines()
        rare_items = str(sheet['F' + str(row)].value).splitlines()

        airship_data.append([sector, wsm, cm, materia, scc, rare_items])

    return airship_data


def export_excel(airship_data):
    # Create blank Workbook
    wb = openpyxl.Workbook()

    # Get active sheet
    sheet = wb.active

    # Sub_Data[0] = Sector (URN kind of)
    # Sub_Data[1][0-7] = Workshop Materials 1-8
    # Sub_Data[2][0-7] = Crafting Materials 1-8
    # Sub_Data[3][0-7] = Materia 1-8 (not useful)
    # Sub_Data[4][0-5] = Shard Crystal Cluster 1-6
    # Sub_Data[5][0-7] = Rare Items 1-8
    sheet['A1'].value = 'Sector'
    sheet['B1'].value = 'Category'
    # Categories = Workshop, Crafting, Materia, Shard Crystal Cluster, Rare
    sheet['C1'].value = 'Item'

    for read, data in enumerate(airship_data):
        row = sheet.max_row + 1

        # Sector
        # sector = sheet['A' + str(row)].value = data[0][1].strip() + ' ' + data[0][2].strip()
        print(data)
        sector = sheet['A' + str(row)].value = data[0][0].strip()

        new_row = 99

        # Workshop Materials
        for item in data[1]:
            if new_row == 99:
                new_row = sheet.max_row
            else:
                new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = sector
            sheet['B' + str(new_row)].value = 'Workshop Materials'
            wsm = sheet.cell(row=new_row, column=3)
            wsm.value = re.sub(r" ?\([^a-z)]+\)", "", item)
        # Crafting Materials
        for item in data[2]:
            new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = sector
            sheet['B' + str(new_row)].value = 'Crafting Materials'
            cm = sheet.cell(row=new_row, column=3)
            cm.value = item
        ''' # Materia
        for item in data[3]:
            new_row = sheet.max_row + 1
            sheet['B' + str(new_row)].value = 'Materia'
            materia = sheet.cell(row=new_row, column=3)
            materia.value = item '''
        # Shard Crystal Cluster
        for item in data[4]:
            new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = sector
            sheet['B' + str(new_row)].value = 'Shard Crystal Cluster'
            scc = sheet.cell(row=new_row, column=3)
            scc.value = item
        # Rare Items
        for item in data[5]:
            new_row = sheet.max_row + 1
            sheet['A' + str(new_row)].value = sector
            sheet['B' + str(new_row)].value = 'Rare Items'
            ri = sheet.cell(row=new_row, column=3)
            ri.value = re.sub(r" ?\([^a-z)]+\)", "", item)

    wb.save('C:\\Users\\jwhitten\\Documents\\Python\\Sub Data\\ExportAirshipData.xlsx')


def main():
    file_path = get_excel()
    airship_data = import_excel(file_path)

    # for x in airship_data:
    # print(x)

    # TODO Export Excel data
    export_excel(airship_data)


if __name__ == "__main__":
    main()
